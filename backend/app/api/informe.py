from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from typing import List, Optional
from decimal import Decimal
from datetime import date
import io

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.user import User
from app.models.obra import (
    Obra, MovimientoCronograma, TipoMovimiento,
    ItemObra, CertificadoAvance, CertificadoItem,
)
from app.models.cliente import Cliente
from app.models.presupuesto import Presupuesto
from app.models.materiales import ListadoMateriales
from app.models.analisis_inversion import AnalisisInversion
from app.schemas.obra import (
    ObraCreate, ObraUpdate, ObraOut,
    MovimientoCreate, MovimientoUpdate, MovimientoOut, AplicarAjusteIPC,
    ResumenCronograma, ImportarCronogramaResultado,
    VincularPresupuesto,
    ItemCreate, ItemUpdate, ItemOut,
    CertificadoCreate, CertificadoOut, CertificadoItemOut,
    ResumenCertificados, CurvaOut, PuntoCurva,
)

router = APIRouter(prefix="/api/clientes/{cliente_id}/obras", tags=["obras"])


def _solo_gaston(user: User):
    if user.rol != "gaston":
        raise HTTPException(status_code=403, detail="Solo Gastón puede acceder a esta sección")


def _orden_natural(orden_str) -> tuple:
    """
    Convierte un código de orden tipo "1", "1.1", "1.1.2" en una tupla
    numérica para poder ordenar de forma jerárquica (no alfabética):
    así "1.2" queda antes que "1.10", por ejemplo.
    """
    if not orden_str:
        return (0,)
    partes = []
    for p in str(orden_str).split("."):
        p = p.strip()
        try:
            partes.append(int(p))
        except ValueError:
            partes.append(0)
    return tuple(partes) if partes else (0,)


def _get_cliente(db: Session, cliente_id: int) -> Cliente:
    c = db.query(Cliente).filter(Cliente.id == cliente_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Cliente no encontrado")
    return c


def _get_obra(db: Session, cliente_id: int, obra_id: int) -> Obra:
    o = db.query(Obra).filter(Obra.id == obra_id, Obra.cliente_id == cliente_id).first()
    if not o:
        raise HTTPException(status_code=404, detail="Obra no encontrada")
    return o


def _get_movimiento(db: Session, obra_id: int, movimiento_id: int) -> MovimientoCronograma:
    m = db.query(MovimientoCronograma).filter(
        MovimientoCronograma.id == movimiento_id, MovimientoCronograma.obra_id == obra_id
    ).first()
    if not m:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")
    return m


def _totales_cronograma(o: Obra) -> dict:
    total_cargos_cliente = sum((m.monto_cliente for m in o.cronograma if m.tipo == TipoMovimiento.cargo), Decimal("0"))
    total_pagos_cliente = sum((m.monto_cliente for m in o.cronograma if m.tipo == TipoMovimiento.pago), Decimal("0"))
    total_cargos_albanil = sum((m.monto_albanil for m in o.cronograma if m.tipo == TipoMovimiento.cargo), Decimal("0"))
    total_pagos_albanil = sum((m.monto_albanil for m in o.cronograma if m.tipo == TipoMovimiento.pago), Decimal("0"))
    return {
        "total_cargos_cliente": total_cargos_cliente,
        "total_pagos_cliente": total_pagos_cliente,
        "saldo_cliente": total_cargos_cliente - total_pagos_cliente,
        "total_cargos_albanil": total_cargos_albanil,
        "total_pagos_albanil": total_pagos_albanil,
        "saldo_albanil": total_cargos_albanil - total_pagos_albanil,
    }


def _enriquecer_obra(o: Obra, es_gaston: bool) -> dict:
    d = {c.name: getattr(o, c.name) for c in o.__table__.columns}
    d["presupuesto_numero"] = o.presupuesto.numero if o.presupuesto else None

    if es_gaston:
        t = _totales_cronograma(o)
        d["total_cliente"] = t["total_cargos_cliente"]
        d["total_albanil"] = t["total_cargos_albanil"]
        d["pagado_cliente"] = t["total_pagos_cliente"]
        d["pagado_albanil"] = t["total_pagos_albanil"]
    else:
        d["total_cliente"] = None
        d["total_albanil"] = None
        d["pagado_cliente"] = None
        d["pagado_albanil"] = None

    return d


def _movimientos_con_saldo(movimientos: list) -> list:
    """Ordena por fecha y calcula el saldo acumulado (cargos - pagos) hasta cada movimiento."""
    ordenados = sorted(movimientos, key=lambda m: (m.fecha, m.id))
    saldo_cliente = Decimal("0")
    saldo_albanil = Decimal("0")
    salida = []
    for m in ordenados:
        signo = 1 if m.tipo == TipoMovimiento.cargo else -1
        saldo_cliente += signo * m.monto_cliente
        saldo_albanil += signo * m.monto_albanil
        d = {c.name: getattr(m, c.name) for c in m.__table__.columns}
        d["saldo_cliente_acumulado"] = saldo_cliente
        d["saldo_albanil_acumulado"] = saldo_albanil
        salida.append(d)
    return salida


# ── Obra: datos generales ────────────────────────────────────────────────────

@router.get("/", response_model=List[ObraOut])
def listar_obras(
    cliente_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _get_cliente(db, cliente_id)
    es_gaston = current_user.rol == "gaston"
    obras = db.query(Obra).filter(Obra.cliente_id == cliente_id).order_by(Obra.created_at.desc()).all()
    return [_enriquecer_obra(o, es_gaston) for o in obras]


@router.post("/", response_model=ObraOut)
def crear_obra(
    cliente_id: int,
    datos: ObraCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _solo_gaston(current_user)
    _get_cliente(db, cliente_id)

    if datos.presupuesto_id:
        presu = db.query(Presupuesto).filter(
            Presupuesto.id == datos.presupuesto_id, Presupuesto.cliente_id == cliente_id
        ).first()
        if not presu:
            raise HTTPException(status_code=404, detail="Presupuesto no encontrado para este cliente")

    o = Obra(cliente_id=cliente_id, **datos.model_dump())
    db.add(o)
    db.commit()
    db.refresh(o)
    return _enriquecer_obra(o, True)


@router.get("/{obra_id}", response_model=ObraOut)
def obtener_obra(
    cliente_id: int,
    obra_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    o = _get_obra(db, cliente_id, obra_id)
    return _enriquecer_obra(o, current_user.rol == "gaston")


@router.put("/{obra_id}", response_model=ObraOut)
def actualizar_obra(
    cliente_id: int,
    obra_id: int,
    datos: ObraUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _solo_gaston(current_user)
    o = _get_obra(db, cliente_id, obra_id)
    for k, v in datos.model_dump(exclude_unset=True).items():
        setattr(o, k, v)
    db.commit()
    db.refresh(o)
    return _enriquecer_obra(o, True)


@router.delete("/{obra_id}")
def eliminar_obra(
    cliente_id: int,
    obra_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _solo_gaston(current_user)
    o = _get_obra(db, cliente_id, obra_id)

    # Los listados de materiales y análisis de inversión son útiles por sí
    # solos: se desvinculan de la obra en vez de borrarse junto con ella.
    db.query(ListadoMateriales).filter(ListadoMateriales.obra_id == o.id).update({"obra_id": None})
    db.query(AnalisisInversion).filter(AnalisisInversion.obra_id == o.id).update({"obra_id": None})

    try:
        db.delete(o)
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"No se pudo eliminar la obra: {str(e)}")

    return {"ok": True}


@router.post("/{obra_id}/vincular-presupuesto", response_model=ObraOut)
def vincular_presupuesto(
    cliente_id: int,
    obra_id: int,
    datos: VincularPresupuesto,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _solo_gaston(current_user)
    o = _get_obra(db, cliente_id, obra_id)
    presu = db.query(Presupuesto).filter(
        Presupuesto.id == datos.presupuesto_id, Presupuesto.cliente_id == cliente_id
    ).first()
    if not presu:
        raise HTTPException(status_code=404, detail="Presupuesto no encontrado para este cliente")
    if presu.estado != "confirmado":
        raise HTTPException(status_code=400, detail="Solo se puede vincular un presupuesto confirmado")
    o.presupuesto_id = presu.id
    db.commit()
    db.refresh(o)
    return _enriquecer_obra(o, True)


@router.get("/{obra_id}/portal-link")
def obtener_link_portal(
    cliente_id: int,
    obra_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Devuelve el token del Portal del Cliente (sección 4.15), generándolo la
    primera vez que se pide. Con ese token se arma el link único (sin
    usuario/contraseña) para enviar por WhatsApp.
    """
    o = _get_obra(db, cliente_id, obra_id)
    if not o.token_portal:
        import uuid
        o.token_portal = uuid.uuid4().hex
        db.commit()
        db.refresh(o)
    return {"token": o.token_portal}


@router.post("/{obra_id}/portal-link/regenerar")
def regenerar_link_portal(
    cliente_id: int,
    obra_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Invalida el link anterior (por si se compartió por error) y genera uno nuevo."""
    _solo_gaston(current_user)
    o = _get_obra(db, cliente_id, obra_id)
    import uuid
    o.token_portal = uuid.uuid4().hex
    db.commit()
    db.refresh(o)
    return {"token": o.token_portal}


# ── Cronograma de pagos ──────────────────────────────────────────────────────

# ── Cronograma de pagos (cuenta corriente) ────────────────────────────────────

@router.get("/{obra_id}/cronograma", response_model=List[MovimientoOut])
def listar_cronograma(
    cliente_id: int,
    obra_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _solo_gaston(current_user)
    o = _get_obra(db, cliente_id, obra_id)
    return _movimientos_con_saldo(o.cronograma)


@router.get("/{obra_id}/cronograma/resumen", response_model=ResumenCronograma)
def resumen_cronograma(
    cliente_id: int,
    obra_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _solo_gaston(current_user)
    o = _get_obra(db, cliente_id, obra_id)
    return _totales_cronograma(o)


@router.post("/{obra_id}/cronograma", response_model=MovimientoOut)
def crear_movimiento(
    cliente_id: int,
    obra_id: int,
    datos: MovimientoCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _solo_gaston(current_user)
    o = _get_obra(db, cliente_id, obra_id)
    m = MovimientoCronograma(obra_id=o.id, **datos.model_dump())
    db.add(m)
    db.commit()
    db.refresh(o)
    return next(x for x in _movimientos_con_saldo(o.cronograma) if x["id"] == m.id)


@router.put("/{obra_id}/cronograma/{movimiento_id}", response_model=MovimientoOut)
def actualizar_movimiento(
    cliente_id: int,
    obra_id: int,
    movimiento_id: int,
    datos: MovimientoUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _solo_gaston(current_user)
    o = _get_obra(db, cliente_id, obra_id)
    m = _get_movimiento(db, obra_id, movimiento_id)
    for k, v in datos.model_dump(exclude_unset=True).items():
        setattr(m, k, v)
    db.commit()
    db.refresh(o)
    return next(x for x in _movimientos_con_saldo(o.cronograma) if x["id"] == m.id)


@router.delete("/{obra_id}/cronograma/{movimiento_id}")
def eliminar_movimiento(
    cliente_id: int,
    obra_id: int,
    movimiento_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _solo_gaston(current_user)
    _get_obra(db, cliente_id, obra_id)
    m = _get_movimiento(db, obra_id, movimiento_id)
    db.delete(m)
    db.commit()
    return {"ok": True}


@router.post("/{obra_id}/cronograma/ajustar-ipc", response_model=MovimientoOut)
def ajustar_ipc(
    cliente_id: int,
    obra_id: int,
    datos: AplicarAjusteIPC,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Aplica el ajuste IPC compuesto sobre el SALDO PENDIENTE TOTAL a la
    fecha indicada (no sobre un movimiento puntual): como el saldo ya
    arrastra los ajustes previos, el resultado es naturalmente compuesto
        ajuste = saldo * (1 + ipc/100) - saldo
    y se registra como un nuevo movimiento tipo "cargo".
    """
    _solo_gaston(current_user)
    o = _get_obra(db, cliente_id, obra_id)

    movs_previos = [m for m in o.cronograma if m.fecha <= datos.fecha]
    saldo_cliente = sum(
        ((m.monto_cliente if m.tipo == TipoMovimiento.cargo else -m.monto_cliente) for m in movs_previos),
        Decimal("0"),
    )
    saldo_albanil = sum(
        ((m.monto_albanil if m.tipo == TipoMovimiento.cargo else -m.monto_albanil) for m in movs_previos),
        Decimal("0"),
    )

    factor = (Decimal("1") + (datos.ipc_pct / Decimal("100")))
    ajuste_cliente = (saldo_cliente * factor - saldo_cliente) if datos.cuenta in ("cliente", "ambas") else Decimal("0")
    ajuste_albanil = (saldo_albanil * factor - saldo_albanil) if datos.cuenta in ("albanil", "ambas") else Decimal("0")

    concepto = f"Ajuste por IPC {datos.ipc_pct}% ({datos.fuente or 'estimado'})"
    m = MovimientoCronograma(
        obra_id=o.id, fecha=datos.fecha, tipo=TipoMovimiento.cargo,
        monto_cliente=ajuste_cliente, monto_albanil=ajuste_albanil,
        concepto=concepto, es_ajuste_ipc=True,
    )
    db.add(m)
    db.commit()
    db.refresh(o)
    return next(x for x in _movimientos_con_saldo(o.cronograma) if x["id"] == m.id)


# ── Importar cronograma desde Excel ───────────────────────────────────────────

_CRON_CAND_FECHA = ["fecha"]
_CRON_CAND_PRESUPUESTO = ["presupuesto", "cargo", "cargos"]
_CRON_CAND_PAGOS = ["pagos", "pago"]
_CRON_CAND_OBS = ["observaciones", "observacion", "observación", "concepto", "detalle"]


@router.post("/{obra_id}/cronograma/importar-excel", response_model=ImportarCronogramaResultado)
async def importar_cronograma_excel(
    cliente_id: int,
    obra_id: int,
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Importa el cronograma desde un Excel tipo cuenta corriente: columnas
    Fecha, Presupuesto (cargo) y/o Pagos, y Observaciones. La columna
    "Resto" se ignora (el sistema la recalcula). Las filas con
    observaciones que mencionen "ipc" o "ajuste" se marcan como ajuste IPC.
    """
    _solo_gaston(current_user)
    o = _get_obra(db, cliente_id, obra_id)

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no está instalado")

    contenido = await archivo.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="No se pudo leer el archivo. ¿Es un .xlsx válido?")

    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        raise HTTPException(status_code=400, detail="El archivo está vacío")

    idx_header = None
    col = {}
    for i, fila in enumerate(filas[:5]):
        normal = [_norm(c) for c in fila]
        col_fecha = next((j for j, v in enumerate(normal) if v in _CRON_CAND_FECHA), None)
        if col_fecha is not None:
            idx_header = i
            col["fecha"] = col_fecha
            col["presupuesto"] = next((j for j, v in enumerate(normal) if v in _CRON_CAND_PRESUPUESTO), None)
            col["pagos"] = next((j for j, v in enumerate(normal) if v in _CRON_CAND_PAGOS), None)
            col["obs"] = next((j for j, v in enumerate(normal) if v in _CRON_CAND_OBS), None)
            break

    if idx_header is None:
        raise HTTPException(status_code=400, detail="No encontré una columna 'Fecha'. Revisá los encabezados del Excel.")
    if col.get("presupuesto") is None and col.get("pagos") is None:
        raise HTTPException(status_code=400, detail="No encontré columnas 'Presupuesto' ni 'Pagos'.")

    def _valor(fila, key):
        idx = col.get(key)
        if idx is None or idx >= len(fila):
            return None
        return fila[idx]

    def _dec_abs(val):
        if val is None:
            return None
        try:
            return abs(Decimal(str(val)))
        except Exception:
            return None

    creados = 0
    omitidas = 0
    avisos = []

    for fila in filas[idx_header + 1:]:
        if fila is None:
            continue
        fecha_val = _valor(fila, "fecha")
        if not fecha_val:
            omitidas += 1
            continue
        try:
            if isinstance(fecha_val, str):
                fecha = date.fromisoformat(fecha_val[:10])
            else:
                fecha = fecha_val.date() if hasattr(fecha_val, "date") else fecha_val
        except Exception:
            omitidas += 1
            avisos.append(f"No pude leer la fecha '{fecha_val}', fila omitida.")
            continue

        obs = _valor(fila, "obs")
        obs_txt = str(obs).strip() if obs else None
        es_ajuste = bool(obs_txt) and ("ipc" in obs_txt.lower() or "ajuste" in obs_txt.lower())

        monto_cargo = _dec_abs(_valor(fila, "presupuesto"))
        monto_pago = _dec_abs(_valor(fila, "pagos"))

        if monto_cargo:
            db.add(MovimientoCronograma(
                obra_id=o.id, fecha=fecha, tipo=TipoMovimiento.cargo,
                monto_cliente=monto_cargo, monto_albanil=Decimal("0"),
                concepto=obs_txt, es_ajuste_ipc=es_ajuste,
            ))
            creados += 1
        if monto_pago:
            db.add(MovimientoCronograma(
                obra_id=o.id, fecha=fecha, tipo=TipoMovimiento.pago,
                monto_cliente=monto_pago, monto_albanil=Decimal("0"),
                concepto=obs_txt, es_ajuste_ipc=False,
            ))
            creados += 1
        if not monto_cargo and not monto_pago:
            omitidas += 1

    if creados > 0:
        avisos.append("Los montos se importaron a la cuenta cliente. La cuenta albañil quedó en $0 — cargala manualmente si corresponde.")

    db.commit()
    return {"movimientos_creados": creados, "filas_omitidas": omitidas, "avisos": avisos}


# ── Ítems del cómputo ─────────────────────────────────────────────────────────

def _item_out(i: ItemObra) -> dict:
    d = {c.name: getattr(i, c.name) for c in i.__table__.columns}
    d["total"] = i.cantidad * i.precio_unitario
    d["total_albanil"] = i.cantidad * i.precio_unitario_albanil
    return d


@router.get("/{obra_id}/items", response_model=List[ItemOut])
def listar_items(
    cliente_id: int,
    obra_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    o = _get_obra(db, cliente_id, obra_id)
    items = db.query(ItemObra).filter(ItemObra.obra_id == o.id).all()
    items.sort(key=lambda i: _orden_natural(i.orden))
    return [_item_out(i) for i in items]


@router.post("/{obra_id}/items", response_model=ItemOut)
def crear_item(
    cliente_id: int,
    obra_id: int,
    datos: ItemCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _solo_gaston(current_user)
    o = _get_obra(db, cliente_id, obra_id)
    i = ItemObra(obra_id=o.id, **datos.model_dump())
    db.add(i)
    db.commit()
    db.refresh(i)
    return _item_out(i)


@router.put("/{obra_id}/items/{item_id}", response_model=ItemOut)
def actualizar_item(
    cliente_id: int,
    obra_id: int,
    item_id: int,
    datos: ItemUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _solo_gaston(current_user)
    _get_obra(db, cliente_id, obra_id)
    i = db.query(ItemObra).filter(ItemObra.id == item_id, ItemObra.obra_id == obra_id).first()
    if not i:
        raise HTTPException(status_code=404, detail="Ítem no encontrado")
    for k, v in datos.model_dump(exclude_unset=True).items():
        setattr(i, k, v)
    db.commit()
    db.refresh(i)
    return _item_out(i)


@router.delete("/{obra_id}/items/{item_id}")
def eliminar_item(
    cliente_id: int,
    obra_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _solo_gaston(current_user)
    _get_obra(db, cliente_id, obra_id)
    i = db.query(ItemObra).filter(ItemObra.id == item_id, ItemObra.obra_id == obra_id).first()
    if not i:
        raise HTTPException(status_code=404, detail="Ítem no encontrado")
    db.delete(i)
    db.commit()
    return {"ok": True}


# ── Importar ítems del cómputo desde Excel ─────────────────────────────────────

_CAND_ORDEN       = ["orden", "item", "nro", "n°", "codigo", "código", "cod", "cod."]
_CAND_DESIGNACION = ["designacion", "designación", "descripcion", "descripción", "detalle", "concepto", "item designacion"]
_CAND_UNIDAD      = ["unidad", "un", "u", "und"]
_CAND_CANTIDAD    = ["cantidad", "cant", "cant.", "qty"]
_CAND_PRECIO_CLIENTE = ["precio cliente", "precio unitario cliente", "precio unitario", "p. unit. cliente", "precio", "p.unit", "precio unit."]
_CAND_PRECIO_ALBANIL = ["precio albañil", "precio albanil", "precio unitario albañil", "precio unitario albanil", "p. unit. albañil", "p. unit. albanil", "precio obrero"]


def _norm(txt) -> str:
    return str(txt).strip().lower() if txt is not None else ""


@router.post("/{obra_id}/items/importar-excel")
async def importar_items_excel(
    cliente_id: int,
    obra_id: int,
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Importa los ítems del cómputo desde un Excel: busca una fila de
    encabezados con columnas de Designación y Cantidad (en cualquier
    orden), y opcionalmente Orden, Unidad, Precio cliente y Precio
    albañil. Crea un ítem por cada fila con datos.
    """
    _solo_gaston(current_user)
    o = _get_obra(db, cliente_id, obra_id)

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no está instalado")

    contenido = await archivo.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="No se pudo leer el archivo. ¿Es un .xlsx válido?")

    ws = wb.active
    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        raise HTTPException(status_code=400, detail="El archivo está vacío")

    idx_header = None
    col = {}
    for i, fila in enumerate(filas[:5]):
        normal = [_norm(c) for c in fila]
        col_desig = next((j for j, v in enumerate(normal) if v in _CAND_DESIGNACION), None)
        col_cant = next((j for j, v in enumerate(normal) if v in _CAND_CANTIDAD), None)
        if col_desig is not None and col_cant is not None:
            idx_header = i
            col["designacion"] = col_desig
            col["cantidad"] = col_cant
            col["orden"] = next((j for j, v in enumerate(normal) if v in _CAND_ORDEN), None)
            col["unidad"] = next((j for j, v in enumerate(normal) if v in _CAND_UNIDAD), None)
            col["precio_cliente"] = next((j for j, v in enumerate(normal) if v in _CAND_PRECIO_CLIENTE), None)
            col["precio_albanil"] = next((j for j, v in enumerate(normal) if v in _CAND_PRECIO_ALBANIL), None)
            break

    if idx_header is None:
        raise HTTPException(
            status_code=400,
            detail="No encontré columnas de Designación y Cantidad en las primeras filas. Revisá los encabezados del Excel.",
        )

    def _valor(fila, key):
        idx = col.get(key)
        if idx is None or idx >= len(fila):
            return None
        return fila[idx]

    creados = 0
    omitidas = 0
    avisos = []

    for fila in filas[idx_header + 1:]:
        if fila is None:
            continue
        desig = _valor(fila, "designacion")
        if not desig or not str(desig).strip():
            omitidas += 1
            continue

        def _dec(val, default=Decimal("0")):
            if val is None:
                return default
            try:
                return Decimal(str(val))
            except Exception:
                return default

        orden_val = _valor(fila, "orden")
        unidad_val = _valor(fila, "unidad")

        i = ItemObra(
            obra_id=o.id,
            orden=str(orden_val).strip() if orden_val is not None else None,
            designacion=str(desig).strip(),
            unidad=str(unidad_val).strip() if unidad_val else None,
            cantidad=_dec(_valor(fila, "cantidad")),
            precio_unitario=_dec(_valor(fila, "precio_cliente")),
            precio_unitario_albanil=_dec(_valor(fila, "precio_albanil")),
        )
        db.add(i)
        creados += 1

    if col.get("precio_cliente") is None:
        avisos.append("No encontré columna de precio para la cuenta cliente — quedó en $0, cargalo manualmente.")
    if col.get("precio_albanil") is None:
        avisos.append("No encontré columna de precio para la cuenta albañil — quedó en $0, cargalo manualmente.")

    db.commit()
    return {"items_creados": creados, "filas_omitidas": omitidas, "avisos": avisos}


# ── Certificados de avance ────────────────────────────────────────────────────

def _certificado_out(cert: CertificadoAvance) -> dict:
    items_out = []
    ejecucion_mes = Decimal("0")
    ejecucion_acum = Decimal("0")
    ejecucion_mes_albanil = Decimal("0")
    ejecucion_acum_albanil = Decimal("0")
    for ci in cert.items:
        items_out.append({
            "id": ci.id,
            "item_id": ci.item_id,
            "designacion": ci.item.designacion if ci.item else None,
            "unidad": ci.item.unidad if ci.item else None,
            "pct_acum_anterior": ci.pct_acum_anterior,
            "pct_acum_nuevo": ci.pct_acum_nuevo,
            "pct_mes": ci.pct_mes,
            "total_item_snapshot": ci.total_item_snapshot,
            "monto_mes": ci.monto_mes,
            "monto_acum": ci.monto_acum,
            "saldo": ci.saldo,
            "total_item_snapshot_albanil": ci.total_item_snapshot_albanil,
            "monto_mes_albanil": ci.monto_mes_albanil,
            "monto_acum_albanil": ci.monto_acum_albanil,
            "saldo_albanil": ci.saldo_albanil,
        })
        ejecucion_mes += ci.monto_mes
        ejecucion_acum += ci.monto_acum
        ejecucion_mes_albanil += ci.monto_mes_albanil
        ejecucion_acum_albanil += ci.monto_acum_albanil

    return {
        "id": cert.id,
        "obra_id": cert.obra_id,
        "numero": cert.numero,
        "periodo": cert.periodo,
        "fecha_certificado": cert.fecha_certificado,
        "created_at": cert.created_at,
        "ejecucion_mes": ejecucion_mes,
        "ejecucion_acum": ejecucion_acum,
        "ejecucion_mes_albanil": ejecucion_mes_albanil,
        "ejecucion_acum_albanil": ejecucion_acum_albanil,
        "items": items_out,
    }


@router.get("/{obra_id}/certificados", response_model=List[CertificadoOut])
def listar_certificados(
    cliente_id: int,
    obra_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    o = _get_obra(db, cliente_id, obra_id)
    certs = (
        db.query(CertificadoAvance)
        .filter(CertificadoAvance.obra_id == o.id)
        .order_by(CertificadoAvance.numero)
        .all()
    )
    return [_certificado_out(c) for c in certs]


@router.post("/{obra_id}/certificados", response_model=CertificadoOut)
def crear_certificado(
    cliente_id: int,
    obra_id: int,
    datos: CertificadoCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Carga el % acumulado de cada ítem para este período. El sistema calcula
    automáticamente el % del mes, $ del mes, $ acumulado y saldo (sección 4.6).
    Lo pueden cargar Gastón, Valentina o Valentín.
    """
    o = _get_obra(db, cliente_id, obra_id)
    if not datos.items:
        raise HTTPException(status_code=400, detail="Cargá el % de al menos un ítem")

    numero = (
        db.query(CertificadoAvance)
        .filter(CertificadoAvance.obra_id == o.id)
        .count()
    ) + 1

    cert = CertificadoAvance(
        obra_id=o.id, numero=numero, periodo=datos.periodo,
        fecha_certificado=datos.fecha_certificado,
    )
    db.add(cert)
    db.flush()  # para tener cert.id antes del commit

    for entrada in datos.items:
        item = db.query(ItemObra).filter(
            ItemObra.id == entrada.item_id, ItemObra.obra_id == o.id
        ).first()
        if not item:
            db.rollback()
            raise HTTPException(status_code=404, detail=f"Ítem {entrada.item_id} no encontrado en esta obra")

        ultimo = (
            db.query(CertificadoItem)
            .join(CertificadoAvance, CertificadoItem.certificado_id == CertificadoAvance.id)
            .filter(CertificadoItem.item_id == item.id, CertificadoAvance.obra_id == o.id)
            .order_by(CertificadoAvance.numero.desc())
            .first()
        )
        pct_acum_anterior = ultimo.pct_acum_nuevo if ultimo else Decimal("0")
        pct_acum_nuevo = entrada.pct_acum_nuevo
        pct_mes = pct_acum_nuevo - pct_acum_anterior

        total_item = item.cantidad * item.precio_unitario
        monto_mes = (pct_mes / Decimal("100")) * total_item
        monto_acum = (pct_acum_nuevo / Decimal("100")) * total_item
        saldo = total_item - monto_acum

        total_item_albanil = item.cantidad * item.precio_unitario_albanil
        monto_mes_albanil = (pct_mes / Decimal("100")) * total_item_albanil
        monto_acum_albanil = (pct_acum_nuevo / Decimal("100")) * total_item_albanil
        saldo_albanil = total_item_albanil - monto_acum_albanil

        db.add(CertificadoItem(
            certificado_id=cert.id, item_id=item.id,
            pct_acum_anterior=pct_acum_anterior, pct_acum_nuevo=pct_acum_nuevo, pct_mes=pct_mes,
            total_item_snapshot=total_item, monto_mes=monto_mes, monto_acum=monto_acum, saldo=saldo,
            total_item_snapshot_albanil=total_item_albanil, monto_mes_albanil=monto_mes_albanil,
            monto_acum_albanil=monto_acum_albanil, saldo_albanil=saldo_albanil,
        ))

    db.commit()
    db.refresh(cert)
    return _certificado_out(cert)


@router.get("/{obra_id}/certificados/{certificado_id}", response_model=CertificadoOut)
def obtener_certificado(
    cliente_id: int,
    obra_id: int,
    certificado_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _get_obra(db, cliente_id, obra_id)
    cert = db.query(CertificadoAvance).filter(
        CertificadoAvance.id == certificado_id, CertificadoAvance.obra_id == obra_id
    ).first()
    if not cert:
        raise HTTPException(status_code=404, detail="Certificado no encontrado")
    return _certificado_out(cert)


@router.delete("/{obra_id}/certificados/{certificado_id}")
def eliminar_certificado(
    cliente_id: int,
    obra_id: int,
    certificado_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _solo_gaston(current_user)
    _get_obra(db, cliente_id, obra_id)
    cert = db.query(CertificadoAvance).filter(
        CertificadoAvance.id == certificado_id, CertificadoAvance.obra_id == obra_id
    ).first()
    if not cert:
        raise HTTPException(status_code=404, detail="Certificado no encontrado")
    db.delete(cert)
    db.commit()
    return {"ok": True}


# ── Resumen y curva ejecutado vs pagos ────────────────────────────────────────

@router.get("/{obra_id}/resumen-certificados", response_model=ResumenCertificados)
def resumen_certificados(
    cliente_id: int,
    obra_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    o = _get_obra(db, cliente_id, obra_id)

    presupuesto_base = sum((i.cantidad * i.precio_unitario for i in o.items_computo), Decimal("0"))
    ajuste_ipc_acumulado = sum(
        (m.monto_cliente for m in o.cronograma if m.es_ajuste_ipc and m.tipo == TipoMovimiento.cargo), Decimal("0")
    )
    total_actualizado = presupuesto_base + ajuste_ipc_acumulado

    presupuesto_base_albanil = sum((i.cantidad * i.precio_unitario_albanil for i in o.items_computo), Decimal("0"))
    ajuste_ipc_acumulado_albanil = sum(
        (m.monto_albanil for m in o.cronograma if m.es_ajuste_ipc and m.tipo == TipoMovimiento.cargo), Decimal("0")
    )
    total_actualizado_albanil = presupuesto_base_albanil + ajuste_ipc_acumulado_albanil

    ultimo_cert = (
        db.query(CertificadoAvance)
        .filter(CertificadoAvance.obra_id == o.id)
        .order_by(CertificadoAvance.numero.desc())
        .first()
    )
    ejecucion_acumulada = Decimal("0")
    ejecucion_acumulada_albanil = Decimal("0")
    if ultimo_cert:
        ejecucion_acumulada = sum((ci.monto_acum for ci in ultimo_cert.items), Decimal("0"))
        ejecucion_acumulada_albanil = sum((ci.monto_acum_albanil for ci in ultimo_cert.items), Decimal("0"))

    saldo_pendiente = total_actualizado - ejecucion_acumulada
    saldo_pendiente_albanil = total_actualizado_albanil - ejecucion_acumulada_albanil

    return {
        "presupuesto_base": presupuesto_base,
        "ajuste_ipc_acumulado": ajuste_ipc_acumulado,
        "total_actualizado": total_actualizado,
        "ejecucion_acumulada": ejecucion_acumulada,
        "saldo_pendiente": saldo_pendiente,
        "presupuesto_base_albanil": presupuesto_base_albanil,
        "ajuste_ipc_acumulado_albanil": ajuste_ipc_acumulado_albanil,
        "total_actualizado_albanil": total_actualizado_albanil,
        "ejecucion_acumulada_albanil": ejecucion_acumulada_albanil,
        "saldo_pendiente_albanil": saldo_pendiente_albanil,
    }


@router.get("/{obra_id}/curva", response_model=CurvaOut)
def curva_ejecutado_vs_pagos(
    cliente_id: int,
    obra_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Serie temporal para graficar Ejecutado (certificados) vs Pagos acumulados
    (cronograma), calculada en paralelo para cuenta cliente y cuenta albañil.
    Alerta si en algún punto los pagos superan lo ejecutado (sección 4.8):
    "El cliente pagó más de lo ejecutado" (o el equivalente para el albañil).
    """
    o = _get_obra(db, cliente_id, obra_id)
    certs = (
        db.query(CertificadoAvance)
        .filter(CertificadoAvance.obra_id == o.id)
        .order_by(CertificadoAvance.numero)
        .all()
    )

    puntos = []
    alerta = False
    alerta_albanil = False
    for cert in certs:
        ejecutado_acum = sum((ci.monto_acum for ci in cert.items), Decimal("0"))
        ejecutado_acum_albanil = sum((ci.monto_acum_albanil for ci in cert.items), Decimal("0"))

        if cert.fecha_certificado:
            pagos_acum = sum(
                (m.monto_cliente for m in o.cronograma if m.tipo == TipoMovimiento.pago and m.fecha <= cert.fecha_certificado),
                Decimal("0"),
            )
            pagos_acum_albanil = sum(
                (m.monto_albanil for m in o.cronograma if m.tipo == TipoMovimiento.pago and m.fecha <= cert.fecha_certificado),
                Decimal("0"),
            )
        else:
            pagos_acum = sum(
                (m.monto_cliente for m in o.cronograma if m.tipo == TipoMovimiento.pago), Decimal("0")
            )
            pagos_acum_albanil = sum(
                (m.monto_albanil for m in o.cronograma if m.tipo == TipoMovimiento.pago), Decimal("0")
            )

        if pagos_acum > ejecutado_acum:
            alerta = True
        if pagos_acum_albanil > ejecutado_acum_albanil:
            alerta_albanil = True

        puntos.append({
            "periodo": cert.periodo,
            "fecha": cert.fecha_certificado,
            "ejecutado_acum": ejecutado_acum,
            "pagos_acum": pagos_acum,
            "ejecutado_acum_albanil": ejecutado_acum_albanil,
            "pagos_acum_albanil": pagos_acum_albanil,
        })

    return {"puntos": puntos, "alerta": alerta, "alerta_albanil": alerta_albanil}
